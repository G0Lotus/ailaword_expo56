import requests
import openpyxl

meeting_id = 61
list_api = 'https://ailaworld.com/service/expo/api/expo_exhibitor/list'
user_get_api = 'https://ailaworld.com/service/expo/api/booth/user_and_video_info'
card_api = 'https://ailaworld.com/service/expo/api/card/info'

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row = 1, column = 1, value = "Company")
ws.cell(row = 1, column = 2, value = "Name")
ws.cell(row = 1, column = 3, value = "Job-title")
ws.cell(row = 1, column = 4, value = "Telephone")
ws.cell(row = 1, column = 5, value = "Wechat")
ws.cell(row = 1, column = 6, value = "Email")

s = requests.Session()
s.headers.update({'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4464.0 Safari/537.36 Edg/91.0.852.0'})


def get_post_json_data(url, data):
    req = s.post(url, data = data)
    return req.json()


def get_user_json_data(url, meeting_booth_id, company_id):
    url = f"{url}?{meeting_booth_id=}&{company_id=}&{meeting_id=}"
    print(f"Get info from {url}")
    req = s.get(url)
    return req.json()



def get_user_name(json_data: dict):
    if json_data.get('user_name_zh', None):
        print(f"Get User Info : {json_data.get('user_name_zh')}")
        return json_data.get('user_name_zh')
    else:
        print(f"Get User Info : {json_data.get('user_name_en')}")
        return json_data.get('user_name_en')


def get_user_job_title(json_data: dict):
    if json_data.get('job_title_zh', None):
        return json_data.get('job_title_zh')
    else:
        return json_data.get('job_title_en')


def get_user_company(json_data: dict):
    if json_data.get('company_name_zh', None):
        return json_data.get('company_name_zh')
    else:
        return json_data.get('company_name_en')


def get_user_telephone(json_data: dict):
    if json_data.get('area_code'):
        if json_data.get('mobile'):
            return f"{json_data.get('area_code')} {json_data.get('mobile')}"
        else:
            return None
    else:
        if json_data.get('mobile'):
            return json_data.get('mobile')
        else:
            return None


def get_user_wechat(json_data: dict):
    return json_data.get('wechat', None)


def get_user_email(json_data: dict):
    return json_data.get('email')


# collect expo users' ID
user_id_list = list()
expo_info = get_post_json_data(list_api, data = 'start=0&limit=999&meeting_id=61&apply_type=1%2C2')
for expo_data in expo_info['data']:
    users_id_data = get_user_json_data(user_get_api, expo_data['booth_id'], expo_data['company_id'])['data']['user_info_list']
    for x in users_id_data:
        user_id_list.append(x['id'])
#clear duplicate user_id
#user_id_list = list(set(user_id_list))    

# collect User Info with user ID
row = 2
for user_id in user_id_list:
    user_info = get_post_json_data(card_api, data = f"{user_id=}")['data']
    ws.cell(row = row, column = 1, value = get_user_company(user_info))
    ws.cell(row = row, column = 2, value = get_user_name(user_info))
    ws.cell(row = row, column = 3, value = get_user_job_title(user_info))
    ws.cell(row = row, column = 4, value = get_user_telephone(user_info))
    ws.cell(row = row, column = 5, value = get_user_wechat(user_info))
    ws.cell(row = row, column = 6, value = get_user_email(user_info))
    row += 1

wb.save('user_info.xlsx')
wb.close()


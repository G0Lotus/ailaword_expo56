import requests
import openpyxl


def get_company(json_data: dict):
    if json_data.get('name_zh'):
        print(f"get {json_data.get('name_zh')}'s info'")
        return json_data.get('name_zh')
    else:
        print(f"get {json_data.get('name_en')}'s info'")
        return json_data.get('name_en')


def get_address(json_data: dict):
    if json_data.get('address_zh'):
        return json_data.get('address_zh')
    else:
        return json_data.get('address_en')


def get_telephone(json_data: dict):
    if json_data.get('area_code'):
        if json_data.get('telephone'):
            return ' '.join([json_data.get('area_code', ''), json_data.get('telephone', '')])
        else:
            return None
    else:
        if json_data.get('telephone'):
            return json_data.get('telephone')
        else:
            return None


def get_email(json_data: dict):
    return json_data.get('email', None)


def get_profile(json_data: dict):
    if json_data.get('profile_zh'):
        return json_data.get('profile_zh')
    else:
        return json_data.get('profile_en')


def get_shipping_line(json_data: dict):
    line = json_data.get('shipping_line', None)
    if line:
        return "\n".join([ f"#{x['code']}({x['name_cn']} | {x['name_en']})" for x in line ])
    else:
        return None


def get_shipping_adventure(json_data: dict):
    adv = json_data.get('business_classification', None)
    if adv:
        adv_list = list()
        for key in adv:
            for id in adv[key]:
                adv_list.append(f"#{id['desc_zh']}")
        return "\n".join(adv_list)
    else:
        return None


def get_shipping_company(json_data: dict):
    company = json_data.get('shipping_company', None)
    if company:
        return "\n".join([ f"#{x['code']}({x['name_cn']} | {x['name_en']})" for x in company ])
    else:
        return None


wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row = 1, column = 1, value = '公司')
ws.cell(row = 1, column = 2, value = '地址')
ws.cell(row = 1, column = 3, value = '电话')
ws.cell(row = 1, column = 4, value = '邮箱')
ws.cell(row = 1, column = 5, value = '简介')
ws.cell(row = 1, column = 6, value = '业务-航线')
ws.cell(row = 1, column = 7, value = '业务-优势')
ws.cell(row = 1, column = 8, value = '业务-船公司')

session = requests.Session()
session.headers.update({'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4456.0 Safari/537.36 Edg/91.0.845.2'})
list_api = 'https://ailaworld.com/service/expo/api/expo_exhibitor/list'
companyinfo_api = 'https://ailaworld.com/service/meeting/api/booth/company_info'


total_company = session.post(list_api, data='start=0&limit=1&meeting_id=61&apply_type=1%2C2').json()['total']
#total_company=10
row = 2

for x in session.post(list_api, data=f'start=0&limit={total_company}&meeting_id=61&apply_type=1%2C2').json()['data']:
    company_id = x['company_id']
    data = session.post(companyinfo_api, data=f'company_id={company_id}&visit_user_id=85062').json()['data']
    ws.cell(row = row, column = 1, value = get_company(data))
    ws.cell(row = row, column = 2, value = get_address(data))
    ws.cell(row = row, column = 3, value = get_telephone(data))
    ws.cell(row = row, column = 4, value = get_email(data))
    ws.cell(row = row, column = 5, value = get_profile(data))
    ws.cell(row = row, column = 6, value = get_shipping_line(data))
    ws.cell(row = row, column = 7, value = get_shipping_adventure(data))
    ws.cell(row = row, column = 8, value = get_shipping_company(data))
    row += 1

wb.save('result.xlsx')
wb.close()
